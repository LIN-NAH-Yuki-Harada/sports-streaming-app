#!/usr/bin/env python3
"""LIVE SPOtCH 月次PL（3年分）Excel生成 — PPT数値統一版"""

import os, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ── スタイル定義 ──
FN = "Meiryo UI"
f_title   = Font(name=FN, size=14, bold=True, color="1E1E2E")
f_section = Font(name=FN, size=11, bold=True, color="FFFFFF")
f_header  = Font(name=FN, size=10, bold=True, color="1E1E2E")
f_item    = Font(name=FN, size=10, color="333333")
f_item_b  = Font(name=FN, size=10, bold=True, color="333333")
f_total   = Font(name=FN, size=11, bold=True, color="1E1E2E")
f_profit  = Font(name=FN, size=11, bold=True, color="00B896")
f_loss    = Font(name=FN, size=11, bold=True, color="EF4444")
f_sub     = Font(name=FN, size=9, color="888899")
f_note    = Font(name=FN, size=9, color="666666")

fill_section = PatternFill("solid", fgColor="2C3E50")
fill_header  = PatternFill("solid", fgColor="E8F5F1")
fill_light   = PatternFill("solid", fgColor="F8F9FB")
fill_white   = PatternFill("solid", fgColor="FFFFFF")
fill_total   = PatternFill("solid", fgColor="E6F9F5")
fill_yr      = PatternFill("solid", fgColor="EFF6FF")

align_c  = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_r  = Alignment(horizontal="right", vertical="center")
align_l  = Alignment(horizontal="left", vertical="center", indent=1)
align_l2 = Alignment(horizontal="left", vertical="center", indent=2)

thin = Side(style="thin", color="E2E8F0")
border_all   = Border(top=thin, bottom=thin, left=thin, right=thin)
thick_top    = Border(top=Side(style="medium", color="1E1E2E"), bottom=thin, left=thin, right=thin)

NUM_FMT     = '#,##0'
NUM_FMT_NEG = '#,##0;[Red]▲#,##0'

# ── 月次ラベル ──
months_label = []
for y, ms in [(2026, range(7, 13)), (2027, range(1, 13)), (2028, range(1, 13)), (2029, range(1, 7))]:
    for m in ms:
        months_label.append(f"{y}/{m}")

# ══════════════════════════════════════════
# 成長モデル（PPTの年次目標に合わせて設計）
# ══════════════════════════════════════════
# 目標:
#   Year 1末: 登録 8,000 / 有料 1,000
#   Year 2末: 登録 35,000 / 有料 4,500
#   Year 3末: 登録 100,000 / 有料 13,000

# 月次新規登録ユーザー（合計が年次目標に一致するよう設計）
growth_y1 = [
    200, 300, 400,   # M1-3: ローンチ直後（夏の大会シーズン）
    500, 600, 500,   # M4-6: 秋の大会シーズン
    400, 500, 800,   # M7-9: 冬〜春（新学期PR強化）
    1000, 1200, 1600, # M10-12: 春の大会シーズン（口コミ加速）
]  # 合計: 8,000

growth_y2 = [
    1800, 2000, 2200,  # M13-15
    2400, 2600, 2200,  # M16-18
    2000, 2200, 2800,  # M19-21
    3000, 3200, 3600,  # M22-24
]  # 合計: 30,000 → 累計 38,000 だが目標35,000なので微調整
# 調整: 合計27,000にする
growth_y2 = [
    1500, 1700, 1900,  # M13-15
    2100, 2300, 2000,  # M16-18
    1800, 2000, 2500,  # M19-21
    2700, 3000, 3500,  # M22-24
]  # 合計: 27,000 → 累計 35,000

growth_y3 = [
    4000, 4500, 5000,  # M25-27
    5500, 5500, 5000,  # M28-30
    5500, 6000, 6500,  # M31-33
    7000, 7500, 8500,  # M34-36
]  # 合計: 75,000 だが目標は累計100,000なので65,000必要
growth_y3 = [
    3500, 4000, 4500,  # M25-27
    5000, 5000, 4500,  # M28-30
    5000, 5500, 6000,  # M31-33
    6500, 7000, 8500,  # M34-36
]  # 合計: 65,000 → 累計 100,000

growth = growth_y1 + growth_y2 + growth_y3

registered = [0] * 36
paid = [0] * 36
team_paid = [0] * 36
individual_paid = [0] * 36

for i in range(36):
    registered[i] = (registered[i - 1] if i > 0 else 0) + growth[i]
    # 有料転換率: 段階的に改善
    if i < 4:
        conv = 0.08
    elif i < 8:
        conv = 0.10
    elif i < 12:
        conv = 0.125   # Year 1末で約1,000人
    elif i < 18:
        conv = 0.12
    elif i < 24:
        conv = 0.13    # Year 2末で約4,500人
    else:
        conv = 0.13    # Year 3末で約13,000人
    paid[i] = int(registered[i] * conv)
    # チームプラン比率
    team_ratio = 0.25 if i < 12 else (0.30 if i < 24 else 0.35)
    team_paid[i] = int(paid[i] * team_ratio)
    individual_paid[i] = paid[i] - team_paid[i]

# ── 月次売上 ──
rev_individual = [individual_paid[i] * 300 for i in range(36)]
rev_team = [team_paid[i] * 500 for i in range(36)]
rev_archive = [0] * 36
for i in range(6, 36):
    rev_archive[i] = int(registered[i] * 0.01 * 100)

# 広告収入: Year 2後半から
rev_ad = [0] * 36
for i in range(18, 24):
    rev_ad[i] = 50000 + (i - 18) * 60000
for i in range(24, 36):
    rev_ad[i] = 500000 + (i - 24) * 100000

rev_total = [rev_individual[i] + rev_team[i] + rev_archive[i] + rev_ad[i] for i in range(36)]

# ── 月次経費 ──
cost_livekit = [5000 + paid[i] * 15 for i in range(36)]

cost_supabase = [0] * 36
for i in range(36):
    r = registered[i]
    cost_supabase[i] = 3000 if r < 5000 else (8000 if r < 20000 else (15000 if r < 50000 else 25000))

cost_vercel = [0] * 36
for i in range(36):
    r = registered[i]
    cost_vercel[i] = 2500 if r < 5000 else (5000 if r < 30000 else 10000)

cost_stripe = [int((rev_individual[i] + rev_team[i]) * 0.036) for i in range(36)]
cost_infra_other = [2000] * 36

cost_infra_total = [cost_livekit[i] + cost_supabase[i] + cost_vercel[i] + cost_stripe[i] + cost_infra_other[i] for i in range(36)]

cost_dev = [0] * 36
for i in range(36):
    cost_dev[i] = 300000 if i < 3 else (200000 if i < 6 else (100000 if i < 12 else (80000 if i < 24 else 150000)))

cost_mktg = [0] * 36
for i in range(36):
    cost_mktg[i] = 300000 if i < 3 else (200000 if i < 6 else (150000 if i < 12 else (200000 if i < 24 else 400000)))

cost_hr = [0] * 36
for i in range(36):
    cost_hr[i] = 0 if i < 12 else (120000 if i < 24 else 350000)

cost_other = [0] * 36
for i in range(36):
    cost_other[i] = 20000 if i < 12 else (30000 if i < 24 else 50000)

# 借入返済（5年60回、利率2.0%想定）
LOAN = 7000000
RATE_M = 0.02 / 12
TERM = 60
MONTHLY_REPAY = int(LOAN * RATE_M / (1 - (1 + RATE_M) ** -TERM))
cost_repay = [MONTHLY_REPAY] * 36  # 3年分（残り2年は計画外）

cost_total = [cost_infra_total[i] + cost_dev[i] + cost_mktg[i] + cost_hr[i] + cost_other[i] for i in range(36)]
cost_total_with_repay = [cost_total[i] + cost_repay[i] for i in range(36)]

profit = [rev_total[i] - cost_total[i] for i in range(36)]
profit_after_repay = [rev_total[i] - cost_total_with_repay[i] for i in range(36)]
cumulative = [0] * 36
cumulative_after = [0] * 36
for i in range(36):
    cumulative[i] = (cumulative[i - 1] if i > 0 else 0) + profit[i]
    cumulative_after[i] = (cumulative_after[i - 1] if i > 0 else 0) + profit_after_repay[i]


# ══════════════════════════════════════════
# シート作成関数
# ══════════════════════════════════════════

def create_sheet(ws, title, start_m, end_m, year_label):
    ws.title = title
    ws.sheet_properties.tabColor = "00B896"

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 3
    from openpyxl.utils import get_column_letter
    for ci in range(3, 16):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions[get_column_letter(15)].width = 16

    ws.merge_cells('A1:O1')
    c = ws['A1']; c.value = f"LIVE SPOtCH  月次損益計算書（{year_label}）"; c.font = f_title
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:O2')
    c = ws['A2']; c.value = "LIN-NAH株式会社 ／ 単位: 円"; c.font = f_sub
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    row = 4
    ws.row_dimensions[row].height = 30
    headers = ["項目", ""] + months_label[start_m:end_m] + ["年間合計"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = f_header; c.fill = fill_header; c.alignment = align_c; c.border = border_all

    sections = [
        ("section", "売上高", None),
        ("item", "配信者プラン（¥300/月）", rev_individual),
        ("item", "チームプラン（¥500/月）", rev_team),
        ("item", "アーカイブ課金", rev_archive),
        ("item", "広告収入", rev_ad),
        ("subtotal", "売上高 合計", rev_total),
        ("blank", None, None),
        ("section", "売上原価・インフラ費", None),
        ("item", "LiveKit Cloud（映像配信）", cost_livekit),
        ("item", "Supabase（DB・認証）", cost_supabase),
        ("item", "Vercel（ホスティング）", cost_vercel),
        ("item", "Stripe手数料（3.6%）", cost_stripe),
        ("item", "その他インフラ", cost_infra_other),
        ("subtotal", "インフラ費 小計", cost_infra_total),
        ("blank", None, None),
        ("section", "販管費", None),
        ("item", "開発費（保守・新機能）", cost_dev),
        ("item", "マーケティング費", cost_mktg),
        ("item", "人件費（CS・運営）", cost_hr),
        ("item", "その他経費", cost_other),
        ("blank", None, None),
        ("total", "経費 合計（返済前）", cost_total),
        ("blank", None, None),
        ("profit", "営業損益（返済前）", profit),
        ("blank", None, None),
        ("section", "借入返済", None),
        ("item", f"月額返済（5年/2.0%/¥{MONTHLY_REPAY:,}）", cost_repay),
        ("blank", None, None),
        ("profit", "返済後キャッシュフロー", profit_after_repay),
        ("cumulative", "累計CF（返済後）", cumulative_after),
        ("blank", None, None),
        ("section", "KPI", None),
        ("kpi", "登録ユーザー（累計）", registered),
        ("kpi", "有料会員（累計）", paid),
        ("kpi", "  うち配信者プラン", individual_paid),
        ("kpi", "  うちチームプラン", team_paid),
    ]

    row = 5
    for stype, label, data in sections:
        if stype == "blank":
            ws.row_dimensions[row].height = 8; row += 1; continue
        ws.row_dimensions[row].height = 24

        if stype == "section":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
            c = ws.cell(row=row, column=1, value=label)
            c.font = f_section; c.fill = fill_section
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(1, 16):
                ws.cell(row=row, column=ci).fill = fill_section
                ws.cell(row=row, column=ci).border = border_all
            row += 1; continue

        c = ws.cell(row=row, column=1, value=label); c.border = border_all

        if stype == "item":
            c.font = f_item; c.alignment = align_l2; bg = fill_light if row % 2 == 0 else fill_white
        elif stype in ("subtotal", "total"):
            c.font = f_total; c.alignment = align_l; bg = fill_total
        elif stype in ("profit", "cumulative"):
            c.font = f_total; c.alignment = align_l; bg = fill_yr
        elif stype == "kpi":
            c.font = f_item; c.alignment = align_l2; bg = fill_light if row % 2 == 0 else fill_white
        c.fill = bg

        ws.cell(row=row, column=2).fill = bg; ws.cell(row=row, column=2).border = border_all

        year_sum = 0
        for mi in range(start_m, end_m):
            ci = 3 + (mi - start_m); val = data[mi]; year_sum += val
            c = ws.cell(row=row, column=ci, value=val)
            c.alignment = align_r; c.border = border_all; c.fill = bg
            if stype in ("profit", "cumulative"):
                c.font = f_profit if val >= 0 else f_loss; c.number_format = NUM_FMT_NEG
            elif stype == "kpi":
                c.font = f_item; c.number_format = '#,##0'
            elif stype in ("subtotal", "total"):
                c.font = f_item_b; c.number_format = NUM_FMT
            else:
                c.font = f_item; c.number_format = NUM_FMT

        ci = 15
        c = ws.cell(row=row, column=ci); c.alignment = align_r; c.fill = bg
        c.border = thick_top if stype == "profit" else border_all
        c.value = data[end_m - 1] if stype in ("cumulative", "kpi") else year_sum
        if stype in ("profit", "cumulative"):
            c.font = Font(name=FN, size=11, bold=True, color="00B896" if c.value >= 0 else "EF4444")
            c.number_format = NUM_FMT_NEG
        elif stype in ("subtotal", "total"):
            c.font = f_total; c.number_format = NUM_FMT
        elif stype == "kpi":
            c.font = f_item_b; c.number_format = '#,##0'
        else:
            c.font = f_item; c.number_format = NUM_FMT
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    ws.cell(row=row, column=1).value = "※ 本計画は見込み数値です。借入条件: 700万円 / 5年返済 / 年利2.0%想定"
    ws.cell(row=row, column=1).font = f_note

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = 'C5'


# ── 3シート作成 ──
ws1 = wb.active
create_sheet(ws1, "Year1（2026.7-2027.6）", 0, 12, "Year 1: 2026年7月〜2027年6月")
ws2 = wb.create_sheet()
create_sheet(ws2, "Year2（2027.7-2028.6）", 12, 24, "Year 2: 2027年7月〜2028年6月")
ws3 = wb.create_sheet()
create_sheet(ws3, "Year3（2028.7-2029.6）", 24, 36, "Year 3: 2028年7月〜2029年6月")


# ══════════════════════════════════════════
# サマリーシート
# ══════════════════════════════════════════
ws0 = wb.create_sheet("サマリー", 0)
ws0.sheet_properties.tabColor = "1E1E2E"
from openpyxl.utils import get_column_letter

ws0.column_dimensions['A'].width = 28
for ci in range(2, 6):
    ws0.column_dimensions[get_column_letter(ci)].width = 18

ws0.merge_cells('A1:E1')
ws0['A1'].value = "LIVE SPOtCH  3年間サマリー"; ws0['A1'].font = f_title
ws0.row_dimensions[1].height = 35
ws0.merge_cells('A2:E2')
ws0['A2'].value = "LIN-NAH株式会社 ／ 単位: 円"; ws0['A2'].font = f_sub
ws0.row_dimensions[2].height = 20; ws0.row_dimensions[3].height = 8

for ci, h in enumerate(["項目", "Year 1", "Year 2", "Year 3", "3年合計"], 1):
    c = ws0.cell(row=4, column=ci, value=h)
    c.font = f_header; c.fill = fill_header; c.alignment = align_c; c.border = border_all
ws0.row_dimensions[4].height = 28

def ysum(d, s, e): return sum(d[s:e])
def ylast(d, e): return d[e - 1]

summary_rows = [
    ("section", "売上高", None, None, None),
    ("item", "配信者プラン", ysum(rev_individual,0,12), ysum(rev_individual,12,24), ysum(rev_individual,24,36)),
    ("item", "チームプラン", ysum(rev_team,0,12), ysum(rev_team,12,24), ysum(rev_team,24,36)),
    ("item", "アーカイブ課金", ysum(rev_archive,0,12), ysum(rev_archive,12,24), ysum(rev_archive,24,36)),
    ("item", "広告収入", ysum(rev_ad,0,12), ysum(rev_ad,12,24), ysum(rev_ad,24,36)),
    ("subtotal", "売上高 合計", ysum(rev_total,0,12), ysum(rev_total,12,24), ysum(rev_total,24,36)),
    ("blank", None, None, None, None),
    ("section", "経費", None, None, None),
    ("item", "インフラ費", ysum(cost_infra_total,0,12), ysum(cost_infra_total,12,24), ysum(cost_infra_total,24,36)),
    ("item", "開発費", ysum(cost_dev,0,12), ysum(cost_dev,12,24), ysum(cost_dev,24,36)),
    ("item", "マーケティング費", ysum(cost_mktg,0,12), ysum(cost_mktg,12,24), ysum(cost_mktg,24,36)),
    ("item", "人件費", ysum(cost_hr,0,12), ysum(cost_hr,12,24), ysum(cost_hr,24,36)),
    ("item", "その他経費", ysum(cost_other,0,12), ysum(cost_other,12,24), ysum(cost_other,24,36)),
    ("subtotal", "経費 合計", ysum(cost_total,0,12), ysum(cost_total,12,24), ysum(cost_total,24,36)),
    ("blank", None, None, None, None),
    ("profit", "営業損益（返済前）", ysum(profit,0,12), ysum(profit,12,24), ysum(profit,24,36)),
    ("blank", None, None, None, None),
    ("section", "借入返済", None, None, None),
    ("item", "年間返済額", MONTHLY_REPAY*12, MONTHLY_REPAY*12, MONTHLY_REPAY*12),
    ("profit", "返済後CF", ysum(profit_after_repay,0,12), ysum(profit_after_repay,12,24), ysum(profit_after_repay,24,36)),
    ("cumulative", "累計CF（返済後）", ylast(cumulative_after,12), ylast(cumulative_after,24), ylast(cumulative_after,36)),
    ("blank", None, None, None, None),
    ("section", "KPI（期末）", None, None, None),
    ("kpi", "登録ユーザー", ylast(registered,12), ylast(registered,24), ylast(registered,36)),
    ("kpi", "有料会員", ylast(paid,12), ylast(paid,24), ylast(paid,36)),
]

row = 5
for stype, label, y1, y2, y3 in summary_rows:
    if stype == "blank":
        ws0.row_dimensions[row].height = 8; row += 1; continue
    ws0.row_dimensions[row].height = 26

    if stype == "section":
        ws0.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws0.cell(row=row, column=1, value=label)
        c.font = f_section; c.fill = fill_section
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for ci in range(1, 6):
            ws0.cell(row=row, column=ci).fill = fill_section
            ws0.cell(row=row, column=ci).border = border_all
        row += 1; continue

    c = ws0.cell(row=row, column=1, value=label); c.border = border_all
    if stype == "item":
        c.font = f_item; c.alignment = align_l2; bg = fill_light if row % 2 == 0 else fill_white
    elif stype == "subtotal":
        c.font = f_total; c.alignment = align_l; bg = fill_total
    elif stype in ("profit", "cumulative"):
        c.font = f_total; c.alignment = align_l; bg = fill_yr
    elif stype == "kpi":
        c.font = f_item; c.alignment = align_l2; bg = fill_light if row % 2 == 0 else fill_white
    c.fill = bg

    vals = [y1, y2, y3]
    total_3y = sum(v for v in vals if v is not None)

    for ci, val in enumerate(vals + [total_3y], 2):
        if val is None: continue
        c = ws0.cell(row=row, column=ci, value=val)
        c.alignment = align_r; c.border = border_all; c.fill = bg; c.number_format = NUM_FMT
        if stype in ("profit", "cumulative"):
            c.font = Font(name=FN, size=11, bold=True, color="00B896" if val >= 0 else "EF4444")
            c.number_format = NUM_FMT_NEG
        elif stype == "subtotal":
            c.font = f_total
        elif stype == "kpi":
            c.font = f_item_b if ci == 5 else f_item
            if ci == 5: c.value = y3
        else:
            c.font = f_item
    row += 1

row += 1
ws0.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws0.cell(row=row, column=1).value = "※ 借入条件: 700万円 / 5年返済 / 年利2.0%想定"
ws0.cell(row=row, column=1).font = f_note
ws0.freeze_panes = 'B5'


# ── 前提条件シート ──
ws_a = wb.create_sheet("前提条件")
ws_a.sheet_properties.tabColor = "888899"
ws_a.column_dimensions['A'].width = 30; ws_a.column_dimensions['B'].width = 40

ws_a.merge_cells('A1:B1')
ws_a['A1'].value = "収支計画の前提条件"; ws_a['A1'].font = f_title
ws_a.row_dimensions[1].height = 30

assumptions = [
    ("", ""),
    ("【料金設定】", ""),
    ("配信者プラン月額", "¥300"),
    ("チームプラン月額", "¥500"),
    ("アーカイブ課金（1ヶ月経過後）", "¥100/回"),
    ("", ""),
    ("【転換率】", ""),
    ("有料転換率（Year 1前半）", "8〜10%"),
    ("有料転換率（Year 1後半）", "10〜12.5%"),
    ("有料転換率（Year 2）", "12〜13%"),
    ("有料転換率（Year 3）", "13%"),
    ("チームプラン比率（Year 1）", "25%"),
    ("チームプラン比率（Year 2）", "30%"),
    ("チームプラン比率（Year 3）", "35%"),
    ("", ""),
    ("【ユーザー目標（PPTと統一）】", ""),
    ("Year 1末 登録ユーザー", "8,000人"),
    ("Year 1末 有料会員", "約1,000人"),
    ("Year 2末 登録ユーザー", "35,000人"),
    ("Year 2末 有料会員", "約4,500人"),
    ("Year 3末 登録ユーザー", "100,000人"),
    ("Year 3末 有料会員", "約13,000人"),
    ("", ""),
    ("【インフラ】", ""),
    ("LiveKit Cloud", "¥5,000/月 + 有料会員×¥15"),
    ("Supabase", "¥3,000〜25,000（段階制）"),
    ("Vercel", "¥2,500〜10,000（段階制）"),
    ("Stripe手数料", "売上の3.6%"),
    ("", ""),
    ("【広告収入】", ""),
    ("広告開始時期", "Year 2後半（19ヶ月目〜）"),
    ("初期広告単価", "¥50,000/月〜"),
    ("", ""),
    ("【借入条件】", ""),
    ("借入金額", "¥7,000,000"),
    ("返済期間", "5年（60回）"),
    ("想定利率", "年2.0%"),
    (f"月額返済額", f"¥{MONTHLY_REPAY:,}"),
    (f"年間返済額", f"¥{MONTHLY_REPAY * 12:,}"),
    ("", ""),
    ("【その他】", ""),
    ("ローンチ時期", "2026年7月"),
    ("黒字化目標", "Year 2中盤（18ヶ月目前後）"),
]

for i, (label, val) in enumerate(assumptions):
    r = i + 3
    c1 = ws_a.cell(row=r, column=1, value=label)
    c2 = ws_a.cell(row=r, column=2, value=val)
    if label.startswith("【"):
        c1.font = Font(name=FN, size=11, bold=True, color="1E1E2E"); c1.fill = fill_header; c2.fill = fill_header
    else:
        c1.font = f_item; c2.font = f_item
    c1.border = border_all; c2.border = border_all


# ── 返済スケジュールシート ──
ws_r = wb.create_sheet("返済スケジュール")
ws_r.sheet_properties.tabColor = "F59E0B"
ws_r.column_dimensions['A'].width = 12
ws_r.column_dimensions['B'].width = 16
ws_r.column_dimensions['C'].width = 16
ws_r.column_dimensions['D'].width = 16
ws_r.column_dimensions['E'].width = 16

ws_r.merge_cells('A1:E1')
ws_r['A1'].value = "借入金返済スケジュール（700万円 / 5年 / 年利2.0%）"
ws_r['A1'].font = f_title; ws_r.row_dimensions[1].height = 30

for ci, h in enumerate(["回", "月額返済", "元金", "利息", "残高"], 1):
    c = ws_r.cell(row=3, column=ci, value=h)
    c.font = f_header; c.fill = fill_header; c.alignment = align_c; c.border = border_all

balance = LOAN
for i in range(60):
    r = 4 + i
    interest = int(balance * RATE_M)
    principal = MONTHLY_REPAY - interest
    if i == 59:
        principal = balance
        interest = MONTHLY_REPAY - principal if MONTHLY_REPAY > principal else 0
    balance = max(0, balance - principal)

    ws_r.cell(row=r, column=1, value=i + 1).font = f_item
    ws_r.cell(row=r, column=1).alignment = align_c
    for ci, val in enumerate([MONTHLY_REPAY, principal, interest, balance], 2):
        c = ws_r.cell(row=r, column=ci, value=val)
        c.font = f_item; c.number_format = NUM_FMT; c.alignment = align_r
    for ci in range(1, 6):
        ws_r.cell(row=r, column=ci).border = border_all
        if i % 2 == 0:
            ws_r.cell(row=r, column=ci).fill = fill_light


# ── 検証値を出力 ──
print(f"=== 数値検証 ===")
print(f"Year 1末 登録: {registered[11]:,}人 / 有料: {paid[11]:,}人")
print(f"Year 2末 登録: {registered[23]:,}人 / 有料: {paid[23]:,}人")
print(f"Year 3末 登録: {registered[35]:,}人 / 有料: {paid[35]:,}人")
print(f"Year 1 売上: {ysum(rev_total,0,12):,}円 ({ysum(rev_total,0,12)/10000:.0f}万円)")
print(f"Year 2 売上: {ysum(rev_total,12,24):,}円 ({ysum(rev_total,12,24)/10000:.0f}万円)")
print(f"Year 3 売上: {ysum(rev_total,24,36):,}円 ({ysum(rev_total,24,36)/10000:.0f}万円)")
print(f"Year 1 経費: {ysum(cost_total,0,12):,}円 ({ysum(cost_total,0,12)/10000:.0f}万円)")
print(f"Year 2 経費: {ysum(cost_total,12,24):,}円 ({ysum(cost_total,12,24)/10000:.0f}万円)")
print(f"Year 3 経費: {ysum(cost_total,24,36):,}円 ({ysum(cost_total,24,36)/10000:.0f}万円)")
print(f"月額返済: ¥{MONTHLY_REPAY:,}")

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "LIVE_SPOtCH_月次PL.xlsx")
wb.save(out)
print(f"\n保存完了: {out}")
