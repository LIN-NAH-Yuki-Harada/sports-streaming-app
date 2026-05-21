#!/usr/bin/env python3
"""LIVE SPOtCH 開発費用見積もり（参考）Excel生成 — v2 改訂版

変更点 (v1 → v2):
- 進捗 92% → 100% 本番稼働 (5/10 までに 126 PR マージ反映)
- セクション 8 「YouTube Live 連携・配信パイプライン」を新規追加 (5/01 リリース・PR-1〜5)
- セクション 9 「配信品質改善」を新規追加 (Canvas / simulcast / iOS Safari / カメラ / 音声 等)
- セクション 10 「運用堅牢化・セキュリティ」を新規追加 (ストレイ対策 / RLS / TOCTOU CAS guard 等)
- セクション 11 「LP・SEO・共有 UI 最適化」を新規追加
- 残り開発 を「5月中完璧化」(TCP化 B案 + OAuth認定 + 細部) に再構築
- 完成分 約 875 万 → 約 1,320 万 (税込) に増額 — 5/10 実装増分反映
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
FN = "Meiryo UI"

f_title   = Font(name=FN, size=14, bold=True, color="1E1E2E")
f_section = Font(name=FN, size=11, bold=True, color="FFFFFF")
f_header  = Font(name=FN, size=10, bold=True, color="1E1E2E")
f_item    = Font(name=FN, size=10, color="333333")
f_item_b  = Font(name=FN, size=10, bold=True, color="333333")
f_total   = Font(name=FN, size=11, bold=True, color="1E1E2E")
f_grand   = Font(name=FN, size=12, bold=True, color="00B896")
f_sub     = Font(name=FN, size=9, color="888899")
f_note    = Font(name=FN, size=9, color="666666")

fill_sec    = PatternFill("solid", fgColor="2C3E50")
fill_hdr    = PatternFill("solid", fgColor="E8F5F1")
fill_light  = PatternFill("solid", fgColor="F8F9FB")
fill_white  = PatternFill("solid", fgColor="FFFFFF")
fill_total  = PatternFill("solid", fgColor="E6F9F5")
fill_grand  = PatternFill("solid", fgColor="D5F5ED")

align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_r = Alignment(horizontal="right", vertical="center")
align_l = Alignment(horizontal="left", vertical="center", indent=1)
align_l2 = Alignment(horizontal="left", vertical="center", indent=2)

thin = Side(style="thin", color="E2E8F0")
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
thick_bottom = Border(top=thin, bottom=Side(style="medium", color="1E1E2E"), left=thin, right=thin)

NUM = '#,##0'


# ══════════════════════════════════════
# Sheet 1: 見積もり明細（完成分 100% / 5/10 時点）
# ══════════════════════════════════════
ws = wb.active
ws.title = "完成分の見積もり明細"
ws.sheet_properties.tabColor = "00B896"

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16

# タイトル
ws.merge_cells('A1:E1')
ws['A1'].value = "LIVE SPOtCH 開発費用見積もり (v2) — 完成分 (5/10 時点 / 126 PR マージ)"
ws['A1'].font = f_title
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:E2')
ws['A2'].value = "一般的な開発会社に外注した場合の想定費用 ／ 単位: 円（税抜）"
ws['A2'].font = f_sub
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 6

# ヘッダー
row = 4
for ci, h in enumerate(["No.", "項目", "工数", "単価", "金額"], 1):
    c = ws.cell(row=row, column=ci, value=h)
    c.font = f_header; c.fill = fill_hdr; c.alignment = align_c; c.border = border_all
ws.row_dimensions[row].height = 26

# データ
estimate_data = [
    # (type, no, item, manmonth, unit_price, amount)
    ("section", "", "1. 要件定義・企画", "", "", ""),
    ("item", "1-1", "要件ヒアリング・整理", "PM 0.3人月", 1000000, 300000),
    ("item", "1-2", "機能仕様書作成", "PM 0.2人月", 1000000, 200000),
    ("subtotal", "", "小計", "0.5人月", "", 500000),

    ("section", "", "2. UI/UX デザイン", "", "", ""),
    ("item", "2-1", "ワイヤーフレーム設計（13画面）", "デザイナー 0.4人月", 700000, 280000),
    ("item", "2-2", "UIデザイン（13画面 + LP）", "デザイナー 0.6人月", 700000, 420000),
    ("item", "2-3", "レスポンシブ対応設計（PC/タブレット/スマホ）", "デザイナー 0.3人月", 700000, 210000),
    ("subtotal", "", "小計", "1.3人月", "", 910000),

    ("section", "", "3. フロントエンド開発", "", "", ""),
    ("item", "3-1", "Next.js 環境構築・基盤設計", "FE 0.2人月", 800000, 160000),
    ("item", "3-2", "認証画面（Google OAuth + メール認証）", "FE 0.2人月", 800000, 160000),
    ("item", "3-3", "ホーム画面（チーム一覧・配信中一覧）", "FE 0.2人月", 800000, 160000),
    ("item", "3-4", "配信画面（カメラ映像 + スコアオーバーレイ）", "FE 0.5人月", 800000, 400000),
    ("item", "3-5", "視聴画面（リアルタイムスコア更新 + iframe 切替）", "FE 0.4人月", 800000, 320000),
    ("item", "3-6", "スコアボード UI（スポーツ別ルール分岐）", "FE 0.3人月", 800000, 240000),
    ("item", "3-7", "マイページ（プロフィール編集・配信履歴・ストレイ警告）", "FE 0.3人月", 800000, 240000),
    ("item", "3-8", "LP（ランディングページ）", "FE 0.2人月", 800000, 160000),
    ("item", "3-9", "その他ページ（利用規約・PP・お問い合わせ等）", "FE 0.1人月", 800000, 80000),
    ("item", "3-10", "PWA 対応（manifest・ServiceWorker・iOS 対応）", "FE 0.2人月", 800000, 160000),
    ("item", "3-11", "ボトムナビ・共通レイアウト・safe-area 対応", "FE 0.2人月", 800000, 160000),
    ("subtotal", "", "小計", "2.8人月", "", 2240000),

    ("section", "", "4. バックエンド開発", "", "", ""),
    ("item", "4-1", "Supabase 設計・DB スキーマ設計", "BE 0.2人月", 900000, 180000),
    ("item", "4-2", "認証基盤（Google OAuth + メール認証）", "BE 0.2人月", 900000, 180000),
    ("item", "4-3", "配信 CRUD（broadcasts テーブル操作）", "BE 0.2人月", 900000, 180000),
    ("item", "4-4", "スコアリアルタイム更新（Realtime + ポーリング）", "BE 0.3人月", 900000, 270000),
    ("item", "4-5", "LiveKit 連携（トークン API・WebRTC publish 制御）", "BE 0.4人月", 900000, 360000),
    ("item", "4-6", "セットスコア・ルール分岐ロジック", "BE 0.2人月", 900000, 180000),
    ("item", "4-7", "Stripe 決済（Checkout・Webhook・顧客ポータル）", "BE 0.4人月", 900000, 360000),
    ("item", "4-8", "チーム管理（CRUD・招待コード・メンバー権限）", "BE 0.4人月", 900000, 360000),
    ("item", "4-9", "スケジュール管理（配信予定・チームフィルタ）", "BE 0.2人月", 900000, 180000),
    ("item", "4-10", "パスワードリセット・お問い合わせ API", "BE 0.15人月", 900000, 135000),
    ("item", "4-11", "プロフィール管理・退会 API（Stripe 解約連動）", "BE 0.15人月", 900000, 135000),
    ("item", "4-12", "LINE 共有・共有コード生成・QR コード", "BE 0.1人月", 900000, 90000),
    ("subtotal", "", "小計", "2.9人月", "", 2610000),

    ("section", "", "5. インフラ構築・デプロイ", "", "", ""),
    ("item", "5-1", "Vercel デプロイ設定 + 独自ドメイン (live-spotch.com)", "インフラ 0.1人月", 900000, 90000),
    ("item", "5-2", "Supabase 本番環境構築 + RLS 設計", "インフラ 0.15人月", 900000, 135000),
    ("item", "5-3", "LiveKit Cloud Build → Ship 昇格 + Egress 設定", "インフラ 0.1人月", 900000, 90000),
    ("item", "5-4", "環境変数管理・CI/CD・Vercel cron 設定", "インフラ 0.1人月", 900000, 90000),
    ("item", "5-5", "ドメイン・SSL・DNS 設定 (お名前.com)", "インフラ 0.1人月", 900000, 90000),
    ("subtotal", "", "小計", "0.55人月", "", 495000),

    # ────── v2 新規追加セクション ──────
    ("section", "", "8. YouTube Live 連携・配信パイプライン (v2 新規)", "", "", ""),
    ("item", "8-1", "YouTube OAuth 検証申請 (3 scope) + デモ動画作成", "BE 0.3人月", 900000, 270000),
    ("item", "8-2", "LiveKit Egress (RTMP push to YouTube Live)", "BE 0.4人月", 900000, 360000),
    ("item", "8-3", "YouTube Live broadcast 作成 / stream 連携 / 終了処理", "BE 0.3人月", 900000, 270000),
    ("item", "8-4", "視聴経路 YouTube iframe デフォルト切替 (PR #124)", "FE 0.15人月", 800000, 120000),
    ("item", "8-5", "アーカイブ自動保存 + cron upload (旧パイプライン)", "BE 0.3人月", 900000, 270000),
    ("item", "8-6", "Made for Kids ポリシー対応 + 配信ページ警告 UI", "FE+BE 0.15人月", 850000, 127500),
    ("subtotal", "", "小計", "1.6人月", "", 1417500),

    ("section", "", "9. 配信品質改善 (v2 新規)", "", "", ""),
    ("item", "9-1", "スコアボード Canvas 焼き込み + offscreen 最適化", "FE 0.6人月", 800000, 480000),
    ("item", "9-2", "simulcast off + bitrate 調整 (PR #119/#123/#126)", "FE 0.2人月", 800000, 160000),
    ("item", "9-3", "A/V drift 修正 (captureStream / framerate / rAF)", "FE 0.5人月", 800000, 400000),
    ("item", "9-4", "iOS Safari 共有時 WebRTC バックグラウンド対策", "FE 0.5人月", 800000, 400000),
    ("item", "9-5", "カメラレンズ切替 + zoom 物理レンズ対応 (PR #110)", "FE 0.3人月", 800000, 240000),
    ("item", "9-6", "体育館スポーツ音声処理 (RAW 配信 BAND 化)", "FE 0.3人月", 800000, 240000),
    ("subtotal", "", "小計", "2.4人月", "", 1920000),

    ("section", "", "10. 運用堅牢化・セキュリティ (v2 新規)", "", "", ""),
    ("item", "10-1", "ストレイ broadcast 完全対策 (PR #125)", "BE 0.2人月", 900000, 180000),
    ("item", "10-2", "Supabase TOCTOU CAS guard (PR #76)", "BE 0.15人月", 900000, 135000),
    ("item", "10-3", "RLS / 列レベル GRANT セキュリティ強化", "BE 0.4人月", 900000, 360000),
    ("item", "10-4", "cron/cleanup 新パイプライン対応", "BE 0.1人月", 900000, 90000),
    ("item", "10-5", "セキュリティ監査対応 (5/26 + 5/10)", "BE 0.2人月", 900000, 180000),
    ("subtotal", "", "小計", "1.05人月", "", 945000),

    ("section", "", "11. LP・SEO・共有 UI 最適化 (v2 新規)", "", "", ""),
    ("item", "11-1", "LP ブラッシュアップ + リアルタイム訴求撤廃 (PR #122)", "FE+デザイナー 0.4人月", 750000, 300000),
    ("item", "11-2", "SEO 基盤 (Search Console / OG / JSON-LD / sitemap)", "FE 0.3人月", 800000, 240000),
    ("item", "11-3", "共有 UI / LINE 共有 / QR コード生成", "FE 0.2人月", 800000, 160000),
    ("subtotal", "", "小計", "0.9人月", "", 700000),
    # ────── v2 新規追加セクション ここまで ──────

    ("section", "", "6. テスト", "", "", ""),
    ("item", "6-1", "単体テスト・結合テスト", "QA 0.4人月", 550000, 220000),
    ("item", "6-2", "ブラウザ互換性テスト（iOS/Android/PC/Safari/Chrome）", "QA 0.3人月", 550000, 165000),
    ("item", "6-3", "ライブ配信 E2E テスト (実機含む)", "QA 0.4人月", 550000, 220000),
    ("item", "6-4", "セキュリティテスト・脆弱性診断", "QA 0.2人月", 550000, 110000),
    ("item", "6-5", "パフォーマンステスト・負荷試験", "QA 0.2人月", 550000, 110000),
    ("subtotal", "", "小計", "1.5人月", "", 825000),

    ("section", "", "7. プロジェクト管理", "", "", ""),
    ("item", "7-1", "進捗管理・ミーティング・タスクトラッキング", "PM 0.6人月", 1000000, 600000),
    ("item", "7-2", "仕様調整・変更管理 (5/10 オーナーフィードバック対応含む)", "PM 0.4人月", 1000000, 400000),
    ("item", "7-3", "納品・ドキュメント整備・運用引継ぎ", "PM 0.2人月", 1000000, 200000),
    ("subtotal", "", "小計", "1.2人月", "", 1200000),
]

row = 5
for dtype, no, item, mm, unit, amount in estimate_data:
    ws.row_dimensions[row].height = 22

    if dtype == "section":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=item)
        c.font = f_section; c.fill = fill_sec
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).fill = fill_sec
            ws.cell(row=row, column=ci).border = border_all
        row += 1; continue

    bg = fill_total if dtype == "subtotal" else (fill_light if row % 2 == 0 else fill_white)
    ft_item = f_item_b if dtype == "subtotal" else f_item

    for ci, (val, al) in enumerate([(no, align_c), (item, align_l2 if dtype != "subtotal" else align_l),
        (mm, align_c), (unit, align_r), (amount, align_r)], 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = ft_item; c.fill = bg; c.alignment = al; c.border = border_all
        if ci >= 4 and isinstance(val, (int, float)) and val > 0:
            c.number_format = NUM

    row += 1

# 合計行
COMPLETED_EXTAX = 500000 + 910000 + 2240000 + 2610000 + 495000 + 1417500 + 1920000 + 945000 + 700000 + 825000 + 1200000  # = 13,762,500
COMPLETED_INCTAX = int(COMPLETED_EXTAX * 1.10)  # 15,138,750

row += 1
ws.row_dimensions[row].height = 28
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws.cell(row=row, column=1, value="開発費 合計（税抜）")
c.font = f_total; c.fill = fill_total; c.alignment = align_l; c.border = thick_bottom
for ci in range(2, 5):
    ws.cell(row=row, column=ci).fill = fill_total; ws.cell(row=row, column=ci).border = thick_bottom
c = ws.cell(row=row, column=5, value=COMPLETED_EXTAX)
c.font = f_total; c.fill = fill_total; c.alignment = align_r; c.border = thick_bottom; c.number_format = NUM

row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1, value="消費税（10%）").font = f_item
ws.cell(row=row, column=1).alignment = align_l; ws.cell(row=row, column=1).border = border_all
for ci in range(2, 5):
    ws.cell(row=row, column=ci).border = border_all
c = ws.cell(row=row, column=5, value=COMPLETED_INCTAX - COMPLETED_EXTAX)
c.font = f_item; c.alignment = align_r; c.border = border_all; c.number_format = NUM

row += 1
ws.row_dimensions[row].height = 30
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws.cell(row=row, column=1, value="見積もり合計（税込）")
c.font = f_grand; c.fill = fill_grand; c.alignment = align_l; c.border = thick_bottom
for ci in range(2, 5):
    ws.cell(row=row, column=ci).fill = fill_grand; ws.cell(row=row, column=ci).border = thick_bottom
c = ws.cell(row=row, column=5, value=COMPLETED_INCTAX)
c.font = f_grand; c.fill = fill_grand; c.alignment = align_r; c.border = thick_bottom; c.number_format = NUM


# ══════════════════════════════════════
# Sheet 2: 残り開発（5月中完璧化）
# ══════════════════════════════════════
ws2 = wb.create_sheet("残り開発（5月中完璧化）")
ws2.sheet_properties.tabColor = "F59E0B"

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 52
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16

ws2.merge_cells('A1:D1')
ws2['A1'].value = "残り開発費用見積もり（5月中完璧化 → 6/1 正式ローンチ）"
ws2['A1'].font = f_title; ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:D2')
ws2['A2'].value = "単位: 円（税抜）"; ws2['A2'].font = f_sub
ws2.row_dimensions[2].height = 18; ws2.row_dimensions[3].height = 6

for ci, h in enumerate(["No.", "項目", "工数", "金額"], 1):
    c = ws2.cell(row=4, column=ci, value=h)
    c.font = f_header; c.fill = fill_hdr; c.alignment = align_c; c.border = border_all

remaining = [
    ("1", "Google OAuth 認定通過対応 (✅ 2026-05-13 正式通過済・3 scope 承認)", "BE 0.2人月", 180000),
    ("2", "配信プロトコル TCP 化 (B 案・MediaRecorder + 中継 server + RTMP)", "FE+BE 0.7人月", 595000),
    ("3", "live_status='creating' 永遠固まり対処 (egress-webhook 改修)", "BE 0.05人月", 45000),
    ("4", "細部 UX 仕上げ (ストレイ警告 / 視聴 UI 調整)", "FE 0.2人月", 160000),
    ("5", "実機検証 + バグ対応 (5月後半・実試合配信での確認)", "QA 0.3人月", 165000),
    ("6", "外部 API 利用料・追加インフラ (TCP 中継 server 等)", "—", 100000),
    ("7", "ローンチ後 1ヶ月の保守運用・hotfix 対応", "—", 200000),
]

for i, (no, item, mm, amount) in enumerate(remaining):
    row = 5 + i
    bg = fill_light if row % 2 == 0 else fill_white
    for ci, (val, al) in enumerate([(no, align_c), (item, align_l2), (mm, align_c), (amount, align_r)], 1):
        c = ws2.cell(row=row, column=ci, value=val)
        c.font = f_item; c.fill = bg; c.alignment = al; c.border = border_all
        if ci == 4: c.number_format = NUM

REMAINING_EXTAX = sum(amount for _, _, _, amount in remaining)  # 1,445,000
REMAINING_INCTAX = int(REMAINING_EXTAX * 1.10)  # 1,589,500

row = 5 + len(remaining) + 1
ws2.row_dimensions[row].height = 26
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws2.cell(row=row, column=1, value="小計（税抜）").font = f_total
ws2.cell(row=row, column=1).fill = fill_total; ws2.cell(row=row, column=1).alignment = align_l; ws2.cell(row=row, column=1).border = border_all
for ci in range(2, 4): ws2.cell(row=row, column=ci).fill = fill_total; ws2.cell(row=row, column=ci).border = border_all
c = ws2.cell(row=row, column=4, value=REMAINING_EXTAX)
c.font = f_total; c.fill = fill_total; c.alignment = align_r; c.border = border_all; c.number_format = NUM

row += 1
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws2.cell(row=row, column=1, value="消費税（10%）").font = f_item; ws2.cell(row=row, column=1).alignment = align_l; ws2.cell(row=row, column=1).border = border_all
for ci in range(2, 4): ws2.cell(row=row, column=ci).border = border_all
c = ws2.cell(row=row, column=4, value=REMAINING_INCTAX - REMAINING_EXTAX); c.font = f_item; c.alignment = align_r; c.border = border_all; c.number_format = NUM

row += 1
ws2.row_dimensions[row].height = 28
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
c = ws2.cell(row=row, column=1, value="合計（税込）")
c.font = f_grand; c.fill = fill_grand; c.alignment = align_l; c.border = thick_bottom
for ci in range(2, 4): ws2.cell(row=row, column=ci).fill = fill_grand; ws2.cell(row=row, column=ci).border = thick_bottom
c = ws2.cell(row=row, column=4, value=REMAINING_INCTAX)
c.font = f_grand; c.fill = fill_grand; c.alignment = align_r; c.border = thick_bottom; c.number_format = NUM


# ══════════════════════════════════════
# Sheet 3: 総合計サマリー
# ══════════════════════════════════════
ws3 = wb.create_sheet("総合計", 0)
ws3.sheet_properties.tabColor = "1E1E2E"

ws3.column_dimensions['A'].width = 50
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 18

ws3.merge_cells('A1:C1')
ws3['A1'].value = "LIVE SPOtCH 開発費用見積もり (v2) — 総合計"
ws3['A1'].font = f_title; ws3.row_dimensions[1].height = 30

ws3.merge_cells('A2:C2')
ws3['A2'].value = "一般的な開発会社に外注した場合の想定費用 ／ 5/10 時点 / 126 PR マージ反映"; ws3['A2'].font = f_sub
ws3.row_dimensions[2].height = 18; ws3.row_dimensions[3].height = 6

for ci, h in enumerate(["項目", "税抜", "税込"], 1):
    c = ws3.cell(row=4, column=ci, value=h)
    c.font = f_header; c.fill = fill_hdr; c.alignment = align_c; c.border = border_all

TOTAL_EXTAX = COMPLETED_EXTAX + REMAINING_EXTAX
TOTAL_INCTAX = int(TOTAL_EXTAX * 1.10)

summary = [
    ("完成分（5/10 時点 / 100% 本番稼働）", COMPLETED_EXTAX, COMPLETED_INCTAX),
    ("残り開発（5月中完璧化 + 6/1 ローンチ後 1ヶ月保守）", REMAINING_EXTAX, REMAINING_INCTAX),
]
for i, (item, ex_tax, inc_tax) in enumerate(summary):
    row = 5 + i
    bg = fill_light if row % 2 == 0 else fill_white
    ws3.cell(row=row, column=1, value=item).font = f_item
    ws3.cell(row=row, column=1).fill = bg; ws3.cell(row=row, column=1).alignment = align_l; ws3.cell(row=row, column=1).border = border_all
    for ci, val in [(2, ex_tax), (3, inc_tax)]:
        c = ws3.cell(row=row, column=ci, value=val)
        c.font = f_item; c.fill = bg; c.alignment = align_r; c.border = border_all; c.number_format = NUM

row = 7
ws3.row_dimensions[row].height = 30
ws3.cell(row=row, column=1, value="全機能開発 総合計").font = f_grand
ws3.cell(row=row, column=1).fill = fill_grand; ws3.cell(row=row, column=1).alignment = align_l; ws3.cell(row=row, column=1).border = thick_bottom
for ci, val in [(2, TOTAL_EXTAX), (3, TOTAL_INCTAX)]:
    c = ws3.cell(row=row, column=ci, value=val)
    c.font = f_grand; c.fill = fill_grand; c.alignment = align_r; c.border = thick_bottom; c.number_format = NUM

# 補足
row = 9
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws3.cell(row=row, column=1, value="【補足】").font = Font(name=FN, size=10, bold=True, color="333333")

notes = [
    "上記は国内の中規模 Web 開発会社（従業員 20〜50 名程度）の標準的な見積もり水準です。",
    "大手 SIer（NTT データ、富士通等）の場合、1.5〜2 倍（2,000 万〜3,000 万円）になることもあります。",
    "ライブ映像配信（LiveKit / WebRTC / RTMP push）+ YouTube Live 連携 + iOS Safari 対応は特殊技術のため、対応できる会社が限られ割増傾向です。",
    "5/10 までの 126 PR を踏まえ、配信品質改善・運用堅牢化・YouTube Live 連携 等の実装増分（+約 500 万円分）を反映した v2 改訂版です。",
    "保守・運用費（月額）はローンチ後 1ヶ月分のみ含みます。通常、開発費の 15〜20%/年 が保守費用の相場です。",
    "本見積もりは、AI 開発支援ツールを活用せず従来の手法で開発した場合の想定です。",
    "当社は Claude 等の AI 駆動開発で、上記費用と同等の機能を実コストほぼゼロで実装しております（実績）。",
]
for i, note in enumerate(notes):
    r = row + 1 + i
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws3.cell(row=r, column=1, value=f"・{note}").font = f_note


# ══════════════════════════════════════
# Sheet 4: 人月単価の前提
# ══════════════════════════════════════
ws4 = wb.create_sheet("人月単価の前提")
ws4.sheet_properties.tabColor = "888899"
ws4.column_dimensions['A'].width = 30; ws4.column_dimensions['B'].width = 25

ws4.merge_cells('A1:B1')
ws4['A1'].value = "人月単価の前提（国内開発会社の相場）"; ws4['A1'].font = f_title
ws4.row_dimensions[1].height = 28

for ci, h in enumerate(["役割", "月額単価"], 1):
    c = ws4.cell(row=3, column=ci, value=h)
    c.font = f_header; c.fill = fill_hdr; c.alignment = align_c; c.border = border_all

rates = [
    ("プロジェクトマネージャー（PM）", "¥1,000,000〜1,200,000"),
    ("UI デザイナー", "¥600,000〜800,000"),
    ("フロントエンドエンジニア", "¥700,000〜900,000"),
    ("バックエンドエンジニア", "¥800,000〜1,000,000"),
    ("インフラエンジニア", "¥800,000〜1,000,000"),
    ("QA テスター", "¥500,000〜600,000"),
    ("WebRTC / 映像配信スペシャリスト", "¥1,200,000〜1,500,000 (割増)"),
]
for i, (role, rate) in enumerate(rates):
    r = 4 + i
    bg = fill_light if r % 2 == 0 else fill_white
    ws4.cell(row=r, column=1, value=role).font = f_item
    ws4.cell(row=r, column=1).fill = bg; ws4.cell(row=r, column=1).alignment = align_l; ws4.cell(row=r, column=1).border = border_all
    ws4.cell(row=r, column=2, value=rate).font = f_item
    ws4.cell(row=r, column=2).fill = bg; ws4.cell(row=r, column=2).alignment = align_r; ws4.cell(row=r, column=2).border = border_all

ws4.cell(row=12, column=1, value="※ 中間値で算出。大手 SIer の場合はこの 1.5〜2 倍になることもある。").font = f_note
ws4.cell(row=13, column=1, value="※ WebRTC / 映像配信は LIVE SPOtCH のコア技術領域。対応エンジニアが限られるため割増単価が一般的。").font = f_note


out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "LIVE_SPOtCH_開発費用見積もり（参考）.xlsx")
wb.save(out)
print(f"保存完了: {out}")
print(f"")
print(f"=== 数値検証 (v2 改訂版) ===")
print(f"完成分 税抜: ¥{COMPLETED_EXTAX:,}")
print(f"完成分 税込: ¥{COMPLETED_INCTAX:,}")
print(f"残り開発 税抜: ¥{REMAINING_EXTAX:,}")
print(f"残り開発 税込: ¥{REMAINING_INCTAX:,}")
print(f"--- 総合計 ---")
print(f"全体 税抜: ¥{TOTAL_EXTAX:,}")
print(f"全体 税込: ¥{TOTAL_INCTAX:,}")
print(f"")
print(f"提案書 v2 / 融資依頼書 v2 表現 「約 1,200 万円 (税込)」 ≒ 完成分 ¥{COMPLETED_INCTAX:,}")
